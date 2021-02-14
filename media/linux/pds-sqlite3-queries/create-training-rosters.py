#!/usr/bin/env python3

# Basic script to create a list of all PDS trainings of a given type.

import sys
sys.path.insert(0, '../../../python')

import os

import logging.handlers
import logging

import ECC
import Google
import PDSChurch
import GoogleAuth
import copy

from datetime import date
from datetime import datetime
from datetime import timedelta

from oauth2client import tools
from apiclient.http import MediaFileUpload

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from pprint import pprint
from pprint import pformat

# Globals

gapp_id         = 'client_id.json'
guser_cred_file = 'user-credentials.json'

now = datetime.now()
us = timedelta(microseconds=now.microsecond)
now = now - us
timestamp = ('{year:04}-{mon:02}-{day:02} {hour:02}:{min:02}'
                .format(year=now.year, mon=now.month, day=now.day,
                        hour=now.hour, min=now.minute))

trainings   = [
    {
        "title"     : 'Communion Minister',
        "gsheet_id" : '1zWsmd5wnyVLwGRjBRX-wEe6xxJL2Bt0NLQAJBdDk4a0',
        'pds_type'  : 'Communion Minister training',
    },
]

def pretty_member(member):
    phones = list()
    key = 'phones'
    if key in member:
        for phone in member[key]:
            if phone['unlisted']:
                continue

            val = '{ph} {type}'.format(ph=phone['number'], type=phone['type'])
            phones.append(val)

    ministries = {'weekday'     :   False,
                  'weekend'     :   False,
                  'homebound'   :   False,}
    key = 'active_ministries'
    if key in member:
        for ministry in member[key]:
            if ministry['Description'] == 'Weekend Communion':
                ministries['weekend'] = True
            if ministry['Description'] == 'Weekday Communion':
                ministries['weekday'] = True
            if ministry['Description'] == 'Homebound Communion':
                ministries['homebound'] = True

    email = PDSChurch.find_any_email(member)[0]

    if (int)(member['family']['ParKey']) > 9000:
        active = "No"
    else:
        active = "Yes"

    m = {
        'mid'       :   member['MemRecNum'],
        'name'      :   member['first']+' '+member['last'],
        'email'     :   email,
        'phones'    :   phones,
        'active'    :   active,
        'weekend'   :   ministries['weekend'],
        'weekday'   :   ministries['weekday'],
        'homebound' :   ministries['homebound']
    }
    
    return m

def find_training(pds_members, training_to_find):
    def _dt_to_int(datetime):
        return 1000000*datetime.month + 10000*datetime.day + datetime.year
    
    out = dict()
    reqcount = 0

    for m in pds_members.values():
        key = 'requirements'
        if key not in m:
            continue

        for req in m[key]:
            if(req['description'] != training_to_find):
                continue
            reqcount += 1
            mem = pretty_member(m)
            sd = req['start_date']
            ed = req['end_date']
            mid = mem['mid']
            if sd not in out:
                out[sd] = dict()
            if mid not in out[sd]:
                out[sd][mid] = list()
            out[sd][mid].append({
                'mid'           :   mid,
                'name'          :   mem['name'],
                'email'         :   mem['email'],
                'phone'         :   mem['phones'][0],
                'start_date'    :   sd,
                'end_date'      :   ed,
                'stage'         :   req['result'],
                'active'        :   mem['active'],
                'weekend'       :   mem['weekend'],
                'weekday'       :   mem['weekday'],
                'homebound'     :   mem['homebound'],
                'note'          :   req['note'],
            })
            print(m['first']+' '+m['last']+": "+str(ed))
    
    active_out = dict()
    for sd in out:
        for mid in out[sd]:
            for line in out[sd][mid]:
                print(line['end_date'])
                if line['end_date'] > now.date():
                    if mid not in active_out:
                        active_out[mid] = {
                            'mid'           :   mid,
                            'name'          :   mem['name'],
                            'email'         :   mem['email'],
                            'phone'         :   mem['phones'][0],
                            'start_date'    :   sd,
                            'end_date'      :   ed,
                            'weekend'       :   mem['weekend'],
                            'weekday'       :   mem['weekday'],
                            'homebound'     :   mem['homebound'],
                            'note'          :   req['note'],
                            }
                    elif line['start_date'] > active_out[mid]['start_date']:
                        active_out[mid] = {
                            'mid'           :   mid,
                            'name'          :   mem['name'],
                            'email'         :   mem['email'],
                            'phone'         :   mem['phones'][0],
                            'start_date'    :   sd,
                            'end_date'      :   ed,
                            'weekend'       :   mem['weekend'],
                            'weekday'       :   mem['weekday'],
                            'homebound'     :   mem['homebound'],
                            'note'          :   req['note'],
                            }
                    else:
                        continue
            else:
                continue
    
    expired_out = dict()
    for sd in out:
        for mid in out[sd]:
            for line in out[sd][mid]: 
                if line['end_date'] < now.date():
                    if mid not in expired_out:
                        expired_out[mid] = {
                            'mid'           :   mid,
                            'name'          :   mem['name'],
                            'email'         :   mem['email'],
                            'phone'         :   mem['phones'][0],
                            'start_date'    :   sd,
                            'end_date'      :   ed,
                            'weekend'       :   mem['weekend'],
                            'weekday'       :   mem['weekday'],
                            'homebound'     :   mem['homebound'],
                            'note'          :   req['note'],
                            }
                    elif line['start_date'] > active_out[mid]['start_date']:
                        active_out[mid] = {
                            'mid'           :   mid,
                            'name'          :   mem['name'],
                            'email'         :   mem['email'],
                            'phone'         :   mem['phones'][0],
                            'start_date'    :   sd,
                            'end_date'      :   ed,
                            'weekend'       :   mem['weekend'],
                            'weekday'       :   mem['weekday'],
                            'homebound'     :   mem['homebound'],
                            'note'          :   req['note'],
                            }
                    else:
                        continue
                else:
                    continue
        
    print(f"Found {reqcount} training records")
    return out, active_out, expired_out

def write_xlsx(all_entries, active_entries, expired_entries, title):
    def _create_everything_columns(row):
        columns = [(f'A{row}', 'Start Date'             ,   30),
                   (f'B{row}', 'End Date'               ,   30),
                   (f'C{row}', 'Member Name'            ,   30),
                   (f'D{row}', 'Email Address'          ,   30),
                   (f'E{row}', 'Phone Number'           ,   50),
                   (f'F{row}', 'Stage of Certification' ,   50),
                   (f'G{row}', 'Active Parishoner?'     ,   50),
                   (f'H{row}', 'Weekend?'               ,   50),
                   (f'I{row}', 'Weekday?'               ,   50),
                   (f'J{row}', 'Homebound?'             ,   50),
                   (f'K{row}', 'Notes'                  ,   50),]
        return columns

    def _create_active_columns(row):
        columns = [(f'A{row}', 'Start Date'             ,   30),
                   (f'B{row}', 'End Date'               ,   30),
                   (f'C{row}', 'Member Name'            ,   30),
                   (f'D{row}', 'Email Address'          ,   30),
                   (f'E{row}', 'Phone Number'           ,   50),
                   (f'F{row}', 'Weekend?'               ,   50),
                   (f'G{row}', 'Weekday?'               ,   50),
                   (f'H{row}', 'Homebound?'             ,   50),
                   (f'I{row}', 'Notes'                  ,   50),]
        return columns

    def _create_expired_columns(row):
        columns = [(f'A{row}', 'Start Date'             ,   30),
                   (f'B{row}', 'End Date'               ,   30),
                   (f'C{row}', 'Member Name'            ,   30),
                   (f'D{row}', 'Email Address'          ,   30),
                   (f'E{row}', 'Phone Number'           ,   50),
                   (f'F{row}', 'Weekend?'               ,   50),
                   (f'G{row}', 'Weekday?'               ,   50),
                   (f'H{row}', 'Homebound?'             ,   50),
                   (f'I{row}', 'Notes'                  ,   50),]
        return columns

    def _write_data_rows(values, sheet_type, row):
        print(f'Writing sheet: {sheet_type}')
        for sd in sorted(values, reverse=True):
            for mid in sorted(values[sd]):
                if sheet_type == 'everything':
                    for entry in values[sd][mid]:
                        col = 1
                        print(f'col = {col}, {type(col)}, row = {row}, {type(row)}')
                        _ = ws.cell(column=col, row=row, value=entry['start_date'])

                        col += 1
                        _ = ws.cell(column=col, row=row, value=entry['end_date'])

                        col += 1
                        _ = ws.cell(column=col, row=row, value=entry['name'])
                        
                        col += 1
                        _ = ws.cell(column=col, row=row, value=entry['email'])

                        col += 1
                        _ = ws.cell(column=col, row=row, value=entry['phone'])
                    
                        if sheet_type == 'everything':
                            col += 1
                            _ = ws.cell(column=col, row=row, value=entry['stage'])

                            col += 1
                            _ = ws.cell(column=col, row=row, value=entry['active'])

                        col +=1
                        _ = ws.cell(column=col, row=row, value=entry['weekend'])

                        col +=1
                        _ = ws.cell(column=col, row=row, value=entry['weekday'])

                        col +=1
                        _ = ws.cell(column=col, row=row, value=entry['homebound'])
                    
                        col += 1
                        _ = ws.cell(column=col, row=row, value=entry['note'])
                    
                        row += 1

    def _create_sheet(sheet_type):
        # Title rows + set column widths
        title_font = Font(color='FFFF00')
        title_fill = PatternFill(fgColor='0000FF', fill_type='solid')
        title_align = Alignment(horizontal='center')

        last_col = 'I'
    
        row = 1
        ws.merge_cells(f'A{row}:{last_col}{row}')
        cell = f'A{row}'
        ws[cell] = f'Training: {title}'
        ws[cell].fill = title_fill
        ws[cell].font = title_font

        row = row + 1
        ws.merge_cells(f'A{row}:{last_col}{row}')
        cell = f'A{row}'
        ws[cell] = f'Last updated: {now}'
        ws[cell].fill = title_fill
        ws[cell].font = title_font

        row = row + 1
        ws.merge_cells(f'A{row}:{last_col}{row}')
        cell = f'A{row}'
        ws[cell] = ''
        ws[cell].fill = title_fill
        ws[cell].font = title_font

        # Freeze the title row
        row = row + 1
        ws.freeze_panes = ws[f'A{row}']

        row = row + 1
        columns = list()
        if(sheet_type == 'everything'):
            columns = _create_everything_columns(row)
        elif(sheet_type == 'active'):
            columns = _create_active_columns(row)
        elif(sheet_type == 'expired'):
            columns = _create_expired_columns(row)

        for cell,value,width in columns:
            ws[cell] = value
            ws[cell].fill = title_fill
            ws[cell].font = title_font
            ws[cell].alignment = title_align
            ws.column_dimensions[cell[0]].width = width

        if(sheet_type == 'everything'):
            _write_data_rows(all_entries, sheet_type, row)
        elif(sheet_type == 'active'):
            _write_data_rows(active_entries, sheet_type, row)
        elif(sheet_type == 'expired'):
            _write_data_rows(expired_entries, sheet_type, row)

    #---------------------------------------------------------------------

    filename = (f'{title} trainings as of {timestamp}.xlsx')

    wb = Workbook()

    ws = wb.create_sheet('Everything')
    _create_sheet('everything')

    ws = wb.create_sheet('Active')
    _create_sheet('active')

    ws = wb.create_sheet('Expired')
    _create_sheet('expired')

    wb.save(filename)
    print(f'Wrote {filename}')

    return filename


#---------------------------------------------------------------------------

def create_roster(pds_members, training, google, log, dry_run):
    # Find training logs
    entries, active_entries, expired_entries = find_training(pds_members=pds_members,
                      training_to_find=training['pds_type'])
    if entries is None or len(entries) == 0:
        print("No trainings of type: {train}".format(train=training['title']))
    
    # Create xlsx file
    filename = write_xlsx(all_entries=entries, active_entries=active_entries, expired_entries=expired_entries, title=training['title'])
    print("Wrote temp XLSX file: {f}".format(f=filename))

    if not dry_run:
        # Upload xlsx to Google
        upload_overwrite(filename=filename, google=google, file_id=training['gsheet_id'],
                     log=log)
        log.debug("Uploaded XLSX file to Google")

        # Remove temp local xlsx file
        try:
            os.unlink(filename)
            log.debug("Unlinked temp XLSX file")
        except:
            log.info("Failed to unlink temp XLSX file!")
            log.error(traceback.format_exc())

#---------------------------------------------------------------------------

def upload_overwrite(filename, google, file_id, log):
    # Strip the trailing ".xlsx" off the Google Sheet name
    gsheet_name = filename
    if gsheet_name.endswith('.xlsx'):
        gsheet_name = gsheet_name[:-5]

    try:
        log.info('Uploading file update to Google file ID "{id}"'
              .format(id=file_id))
        metadata = {
            'name'     : gsheet_name,
            'mimeType' : Google.mime_types['sheet'],
            'supportsAllDrives' : True,
            }
        media = MediaFileUpload(filename,
                                mimetype=Google.mime_types['sheet'],
                                resumable=True)
        file = google.files().update(body=metadata,
                                     fileId=file_id,
                                     media_body=media,
                                     supportsAllDrives=True,
                                     fields='id').execute()
        log.debug('Successfully updated file: "{filename}" (ID: {id})'
              .format(filename=filename, id=file['id']))

    except:
        log.error('Google file update failed for some reason:')
        log.error(traceback.format_exc())
        exit(1)

#------------------------------------------------------------------
        
def setup_cli_args():
    tools.argparser.add_argument('--logfile',
                                 help='Also save to a logfile')

    tools.argparser.add_argument('--debug',
                                 action='store_true',
                                 default=False,
                                 help='Be extra verbose')

    tools.argparser.add_argument('--dry_run',
                                 action='store_true',
                                 default=False,
                                 help='Do not upload to Google')

    tools.argparser.add_argument('--sqlite3-db',
                                 required=True,
                                 help='Location of PDS sqlite3 database')

    global gapp_id
    tools.argparser.add_argument('--app-id',
                                 default=gapp_id,
                                 help='Filename containing Google application credentials')
    global guser_cred_file
    tools.argparser.add_argument('--user-credentials',
                                 default=guser_cred_file,
                                 help='Filename containing Google user credentials')
    
    args = tools.argparser.parse_args()

    return args

#-------------------------------------------------------------------

def main():

    args = setup_cli_args()

    log = ECC.setup_logging(info=True,
                            debug=args.debug,
                            logfile=args.logfile)

    log.info("Reading PDS data...")
    (pds, pds_families,
     pds_members) = PDSChurch.load_families_and_members(filename=args.sqlite3_db,
                                                        parishioners_only=False,
                                                        log=log)

    apis = {
        'drive' : { 'scope'       : Google.scopes['drive'],
                    'api_name'    : 'drive',
                    'api_version' : 'v3', },
    }
    if not args.dry_run:
        services = GoogleAuth.service_oauth_login(apis,
                                                app_json=args.app_id,
                                                user_json=args.user_credentials,
                                                log=log)
        google = services['drive']
    
    if args.dry_run:
        google = ''

    for training in trainings:
        create_roster(pds_members=pds_members,
                      training=training,
                      google=google,
                      log=log,
                      dry_run=args.dry_run)

    # All done
    pds.connection.close()

main()
